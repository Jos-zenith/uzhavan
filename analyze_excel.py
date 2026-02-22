#!/usr/bin/env python3
import json
from openpyxl import load_workbook
from pathlib import Path

# Load the workbook
excel_path = Path(r'my-app/data/uzhavan.xlsx')
wb = load_workbook(excel_path)

# Get all sheets
all_sheets = wb.sheetnames
print(f"ALL SHEETS: {json.dumps(all_sheets, indent=2)}\n")

# Function to analyze a sheet
def analyze_sheet(sheet_name, max_data_rows=5):
    ws = wb[sheet_name]
    result = {
        "sheet_name": sheet_name,
        "total_rows": ws.max_row,
        "total_columns": ws.max_column,
        "headers": [],
        "sample_rows": [],
        "column_types": []
    }
    
    # Get headers - look for the first row with actual content
    header_row_idx = 1
    for row_idx in range(1, min(10, ws.max_row + 1)):
        row_data = [ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]
        if any(v is not None for v in row_data):
            result["headers"] = row_data
            header_row_idx = row_idx
            break
    
    # Collect sample rows starting from row after headers
    for row_idx in range(header_row_idx + 1, min(header_row_idx + 1 + max_data_rows, ws.max_row + 1)):
        row_data = [ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]
        if any(v is not None for v in row_data):
            result["sample_rows"].append(row_data)
    
    # Determine column types
    for col_idx in range(1, ws.max_column + 1):
        col_name = ws.cell(header_row_idx, col_idx).value or f"Column_{col_idx}"
        col_types = set()
        for row_idx in range(header_row_idx + 1, min(header_row_idx + 20, ws.max_row + 1)):
            cell_val = ws.cell(row_idx, col_idx).value
            if cell_val is not None:
                col_types.add(type(cell_val).__name__)
        result["column_types"].append({
            "column": col_name,
            "types": list(col_types) if col_types else ["Empty/None"]
        })
    
    return result

# Analyze MSME sheet
print("="*80)
print("ANALYZING MSME SHEET")
print("="*80)
msme_info = analyze_sheet("msme", max_data_rows=3)
print(json.dumps(msme_info, indent=2, default=str))

print("\n" + "="*80)
print("ANALYZING CHARTER SHEET")
print("="*80)
charter_info = analyze_sheet("charter", max_data_rows=3)
print(json.dumps(charter_info, indent=2, default=str))

# Create final JSON response
response = {
    "sheets": all_sheets,
    "msme_sheet": {
        "columns": [
            {
                "name": msme_info["headers"][i] if i < len(msme_info["headers"]) else f"Column_{i+1}",
                "data_type": msme_info["column_types"][i]["types"][0] if i < len(msme_info["column_types"]) and msme_info["column_types"][i]["types"] else "Unknown",
                "description": ""
            }
            for i in range(len(msme_info["headers"]))
        ],
        "sample_rows": msme_info["sample_rows"][:3],
        "total_rows": msme_info["total_rows"],
        "key_purpose": "Information about incentive schemes for MSMEs in Tamil Nadu"
    },
    "charter_sheet": {
        "columns": [
            {
                "name": charter_info["headers"][i] if i < len(charter_info["headers"]) else f"Column_{i+1}",
                "data_type": charter_info["column_types"][i]["types"][0] if i < len(charter_info["column_types"]) and charter_info["column_types"][i]["types"] else "Unknown",
                "description": ""
            }
            for i in range(len(charter_info["headers"]))
        ],
        "sample_rows": charter_info["sample_rows"][:3],
        "total_rows": charter_info["total_rows"],
        "key_purpose": "Government schemes and programs charter information"
    }
}

print("\n" + "="*80)
print("FINAL JSON RESPONSE")
print("="*80)
print(json.dumps(response, indent=2, default=str))
