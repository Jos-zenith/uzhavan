#!/usr/bin/env python3
import json
from openpyxl import load_workbook
from pathlib import Path

# Load the workbook
excel_path = Path(r'my-app/data/uzhavan.xlsx')
wb = load_workbook(excel_path)

# Get all sheets
all_sheets = wb.sheetnames

# ============================================
# MSME Sheet Analysis
# ============================================
ws_msme = wb["msme"]

# From inspection: Row 1 = Title, Row 2 = Headers, Row 3+ = Data
msme_title = ws_msme.cell(1, 1).value
msme_header_row = 2
msme_data_start = 3

# Get actual headers from row 2
msme_headers = []
for col in range(1, ws_msme.max_column + 1):
    header_val = ws_msme.cell(msme_header_row, col).value
    if header_val is not None:
        msme_headers.append(header_val)

# Get sample rows (rows 3, 5, 10 to show variety)
msme_samples = []
for row_idx in [3, 5, 10]:
    if row_idx <= ws_msme.max_row:
        row_data = []
        for col in range(1, len(msme_headers) + 1):
            val = ws_msme.cell(row_idx, col).value
            row_data.append(val)
        if any(v is not None for v in row_data):
            msme_samples.append(row_data)

# Determine number of actual data rows (count non-empty rows after header)
msme_data_rows = 0
for row_idx in range(msme_data_start, ws_msme.max_row + 1):
    row_data = [ws_msme.cell(row_idx, col).value for col in range(1, ws_msme.max_column + 1)]
    if any(v is not None for v in row_data):
        msme_data_rows += 1

# ============================================
# CHARTER Sheet Analysis
# ============================================
ws_charter = wb["charter"]

# From inspection: Row 1 = Title, Row 2 = Description, Row 3 = Headers, Row 4+ = Data
charter_title = ws_charter.cell(1, 1).value

# Charter structure - Column A has most data, columns C and D have secondary data
# The main headers appear to be in row 3
charter_header_row = 3
charter_data_start = 4

# Get headers from row 3
charter_headers = []
for col in range(1, ws_charter.max_column + 1):
    header_val = ws_charter.cell(charter_header_row, col).value
    if header_val is not None:
        charter_headers.append(header_val)

# Get sample rows
charter_samples = []
for row_idx in [4, 6, 8]:
    if row_idx <= ws_charter.max_row:
        row_data = []
        for col in range(1, len(charter_headers) + 1):
            val = ws_charter.cell(row_idx, col).value
            row_data.append(val)
        if any(v is not None for v in row_data):
            charter_samples.append(row_data)

# Count data rows
charter_data_rows = 0
for row_idx in range(charter_data_start, ws_charter.max_row + 1):
    row_data = [ws_charter.cell(row_idx, col).value for col in range(1, ws_charter.max_column + 1)]
    if any(v is not None for v in row_data):
        charter_data_rows += 1

# ============================================
# Build JSON Response
# ============================================

response = {
    "sheets": all_sheets,
    "msme_sheet": {
        "sheet_name": "msme",
        "title": msme_title,
        "columns": [
            {
                "name": msme_headers[i] if i < len(msme_headers) else f"Column_{i+1}",
                "data_type": "string (with various content types)",
                "description": "Each column represents a different aspect of the MSME incentive scheme"
            }
            for i in range(len(msme_headers))
        ],
        "sample_rows": msme_samples[:2],  # First 2 data row samples
        "total_rows": msme_data_rows,
        "total_columns": len(msme_headers),
        "key_purpose": "Contains information about incentive schemes available to Micro, Small and Medium Enterprises in Tamil Nadu, including infrastructure support, capital subsidies, employment schemes, and eligibility criteria"
    },
    "charter_sheet": {
        "sheet_name": "charter",
        "title": charter_title,
        "columns": [
            {
                "name": charter_headers[i] if i < len(charter_headers) else f"Column_{i+1}",
                "data_type": "string (text descriptions)",
                "description": f"Column {i+1} of charter sheet"
            }
            for i in range(len(charter_headers))
        ],
        "sample_rows": charter_samples[:2],  # First 2 data row samples
        "total_rows": charter_data_rows,
        "total_columns": len(charter_headers),
        "key_purpose": "Contains a charter/directory of government schemes and programs organized by department (Agriculture, Cooperatives, etc.) including welfare schemes, eligibility conditions, and officer contacts for farmers and other beneficiaries"
    },
    "relationships": {
        "description": "Both sheets document government schemes and benefits",
        "msme_focus": "Financial incentives for enterprise development",
        "charter_focus": "Welfare schemes and eligibility for farmer benefits",
        "common_purpose": "Support economic development and farmer welfare in Tamil Nadu"
    }
}

print(json.dumps(response, indent=2))
