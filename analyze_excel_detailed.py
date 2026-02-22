#!/usr/bin/env python3
import json
from openpyxl import load_workbook
from pathlib import Path

# Load the workbook
excel_path = Path(r'my-app/data/uzhavan.xlsx')
wb = load_workbook(excel_path)

# Get all sheets
all_sheets = wb.sheetnames

# Function to analyze a sheet more intelligently
def analyze_sheet_detailed(sheet_name):
    ws = wb[sheet_name]
    result = {
        "sheet_name": sheet_name,
        "total_rows": ws.max_row,
        "total_columns": ws.max_column,
        "row_breakdown": []
    }
    
    # Scan first 20 rows to understand structure
    for row_idx in range(1, min(21, ws.max_row + 1)):
        row_data = [ws.cell(row_idx, col).value for col in range(1, min(ws.max_column + 1, 11))]
        # Count non-None values
        non_none_count = sum(1 for v in row_data if v is not None)
        result["row_breakdown"].append({
            "row": row_idx,
            "non_empty_cols": non_none_count,
            "data_sample": [str(v)[:40] if v is not None else None for v in row_data]
        })
    
    return result

print("="*100)
print("DETAILED ANALYSIS OF MSME SHEET")
print("="*100)
msme_detailed = analyze_sheet_detailed("msme")
print(json.dumps(msme_detailed, indent=2))

print("\n" + "="*100)
print("DETAILED ANALYSIS OF CHARTER SHEET")
print("="*100)
charter_detailed = analyze_sheet_detailed("charter")
print(json.dumps(charter_detailed, indent=2))

# Now extract real structure
print("\n" + "="*100)
print("EXTRACTING REAL STRUCTURE - MSME")
print("="*100)

ws_msme = wb["msme"]

# Row 1 is the title, Row 2 should have the actual headers
actual_headers_msme = [ws_msme.cell(2, col).value for col in range(1, ws_msme.max_column + 1)]
print("Headers from Row 2:", actual_headers_msme)

# Get sample data rows (starting from row 3)
print("\nSample data rows (rows 3-7):")
for row_idx in range(3, min(8, ws_msme.max_row + 1)):
    row_data = [ws_msme.cell(row_idx, col).value for col in range(1, ws_msme.max_column + 1)]
    print(f"Row {row_idx}: {row_data[:6]}...")  # Show first 6 columns

print("\n" + "="*100)
print("EXTRACTING REAL STRUCTURE - CHARTER")
print("="*100)

ws_charter = wb["charter"]

# Charter sheet structure - scan to find actual headers
print("First 10 rows of charter sheet:")
for row_idx in range(1, min(11, ws_charter.max_row + 1)):
    row_data = [ws_charter.cell(row_idx, col).value for col in range(1, min(ws_charter.max_column + 1, 5))]
    non_empty = [v for v in row_data if v is not None]
    print(f"Row {row_idx}: {non_empty[:3]}...")

# Get more comprehensive view
print("\n" + "="*100)
print("COMPARING BOTH SHEETS - KEY CHARACTERISTICS")
print("="*100)

summary = {
    "sheets": all_sheets,
    "msme": {
        "name": "msme",
        "total_rows": ws_msme.max_row,
        "total_columns": ws_msme.max_column,
        "title_row": "INCENTIVE SCHEMES AVAILABLE TO MICRO, SMALL AND MEDIUM ENTERPRISES IN TAMIL NADU",
        "actual_header_row": 2,
        "data_start_row": 3,
        "actual_column_count_with_data": sum(1 for col in range(1, ws_msme.max_column + 1) if any(ws_msme.cell(row, col).value for row in range(3, min(10, ws_msme.max_row))))
    },
    "charter": {
        "name": "charter",
        "total_rows": ws_charter.max_row,
        "total_columns": ws_charter.max_column,
        "title_row": "1. AGRICULTURE DEPARTMENT",
        "data_start_row": 3,
        "actual_column_count_with_data": sum(1 for col in range(1, ws_charter.max_column + 1) if any(ws_charter.cell(row, col).value for row in range(3, min(10, ws_charter.max_row))))
    }
}

print(json.dumps(summary, indent=2))
